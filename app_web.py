"""
=============================================================================
ANALIZADOR LEGAL — INTERFAZ WEB
=============================================================================
Ejecuta:  python app_web.py  |  Abre: http://localhost:5000
Variables de entorno opcionales:
  APP_PASSWORD   -> Contraseña de acceso (vacío = sin login)
  SECRET_KEY     -> Clave de sesión Flask
  GEMINI_API_KEY -> Si se configura, el campo API Key se oculta en la web
=============================================================================
"""

import os, sys, base64, csv, re, secrets
from pathlib import Path
from datetime import datetime
from functools import wraps
from flask import (Flask, request, jsonify, send_file,
                   render_template_string, session, redirect, url_for)

# ── Dependencias ─────────────────────────────────────────────────────────────
FALTANTES = []
try:
    import google.generativeai as genai
except: FALTANTES.append("google-generativeai")
try:    import fitz
except: FALTANTES.append("PyMuPDF")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except: FALTANTES.append("openpyxl")
try:
    from pdf2image import convert_from_path
    PDF2IMAGE_OK = True
except:
    PDF2IMAGE_OK = False
try:
    from docx import Document as DocxDoc
    DOCX_OK = True
except:
    DOCX_OK = False

if FALTANTES:
    print(f"\n❌ Instala: pip install {' '.join(FALTANTES)}")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
app            = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
UPLOAD_DIR     = Path("uploads_web"); UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR     = Path("resultados_legales"); OUTPUT_DIR.mkdir(exist_ok=True)
MAX_PAGINAS    = 30
MODELO         = "gemini-2.0-flash"
APP_PASSWORD   = os.environ.get("APP_PASSWORD", "")
API_KEY_ENV    = os.environ.get("GEMINI_API_KEY", "")   # Si está, se oculta el campo en la web

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if APP_PASSWORD and not session.get("autenticado"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ── Prompt legal ──────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """Eres un experto en Derecho Corporativo Mexicano con especialización en revisión de Libros Sociales y Actas Constitutivas, con pericia en auditoría de cumplimiento para el registro REPSE.

Tu tarea es analizar el documento adjunto y extraer la información necesaria para el Registro Maestro de la sociedad.

INSTRUCCIONES:
1. Determina el tipo de documento (Constitución, Asamblea Ordinaria, Extraordinaria o Sesión de Consejo).
2. Localiza las cláusulas que impacten los rubros de la tabla.
3. Si es una modificación, identifica el estado anterior y el nuevo estado.
4. Genera ESTRICTAMENTE una tabla Markdown: | Categoría | Dato Actualizado / Cláusula | Observaciones / Cambios Clave |

CAMPOS OBLIGATORIOS:
- Tipo de Documento
- Denominación Social
- Domicilio Social
- Objeto Social
- Capital Social (fijo, variable y total)
- Generales de los Socios (entrada/salida)
- Tenencia Accionaria (por socio)
- Órganos de Administración
- Facultades / Poderes
- Comisario
- Duración de la Sociedad
- Fecha del Acta
- Notaría / Fedatario
- Observaciones REPSE

RESTRICCIONES:
- NO inventes datos. Si un campo no aparece: "Sin cambios en este documento"
- Si sección ilegible: "Sección ilegible - verificar documento original"
- Responde SOLO con la tabla Markdown, sin texto adicional."""

# ── Extracción ────────────────────────────────────────────────────────────────
def extraer_texto_pdf(ruta):
    doc = fitz.open(str(ruta))
    texto = "".join(doc[i].get_text() for i in range(min(len(doc), MAX_PAGINAS)))
    doc.close()
    return texto.strip(), len(texto.strip()) < 100

def pdf_a_imagenes(ruta):
    if not PDF2IMAGE_OK: return []
    paginas = convert_from_path(str(ruta), dpi=200, first_page=1, last_page=MAX_PAGINAS)
    imgs = []
    for i, p in enumerate(paginas):
        tmp = UPLOAD_DIR / f"_tmp_{i}.png"
        p.save(str(tmp), "PNG")
        with open(tmp, "rb") as f:
            data = f.read()
        imgs.append({"mime_type": "image/png", "data": base64.b64encode(data).decode()})
        tmp.unlink(missing_ok=True)
    return imgs

def extraer_texto_docx(ruta):
    if not DOCX_OK: return ""
    doc = DocxDoc(str(ruta))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

def llamar_gemini(api_key, texto="", imagenes_bytes=None):
    genai.configure(api_key=api_key)
    modelo = genai.GenerativeModel(
        model_name=MODELO,
        system_instruction=SYSTEM_PROMPT
    )
    if imagenes_bytes:
        partes = imagenes_bytes + ["Analiza estas páginas del acta y genera la tabla Markdown con TODOS los campos obligatorios."]
    else:
        partes = [f"Analiza este documento legal y genera la tabla:\n\n---\n{texto[:150000]}\n---"]
    r = modelo.generate_content(partes)
    return r.text

def parsear_tabla(md):
    filas = []
    for linea in md.splitlines():
        linea = linea.strip()
        if not linea.startswith("|"): continue
        if re.match(r"^\|[-\s|]+\|$", linea): continue
        if "Categoría" in linea or "Categoria" in linea: continue
        celdas = [c.strip() for c in linea.split("|") if c.strip()]
        if len(celdas) >= 2:
            filas.append({"cat":celdas[0],"dato":celdas[1],"obs":celdas[2] if len(celdas)>2 else ""})
    return filas

def procesar_archivo(ruta, api_key):
    ext = ruta.suffix.lower()
    texto, es_escaneo, imagenes = "", False, []
    if ext == ".pdf":
        texto, es_escaneo = extraer_texto_pdf(ruta)
        if es_escaneo:
            imagenes = pdf_a_imagenes(ruta)
            if not imagenes:
                raise ValueError("PDF escaneado sin soporte de Poppler. Verifica el servidor.")
    elif ext == ".docx":
        texto = extraer_texto_docx(ruta)
        if not texto.strip():
            raise ValueError("El DOCX está vacío o no se pudo leer.")
    else:
        raise ValueError(f"Formato no soportado: {ext}")

    # Gemini acepta imágenes como partes inline
    imagenes_gemini = None
    if es_escaneo and imagenes:
        imagenes_gemini = [{"inline_data": img} for img in imagenes]

    tabla_md = llamar_gemini(api_key, texto=texto, imagenes_bytes=imagenes_gemini)
    filas = parsear_tabla(tabla_md)
    if not filas:
        raise ValueError("Gemini no devolvió datos estructurados. Verifica la calidad del PDF.")
    tipo = next((f["dato"] for f in filas if "Tipo" in f["cat"]), "No identificado")
    return filas, tipo

# ── Excel (lote) ──────────────────────────────────────────────────────────────
def guardar_excel_lote(resultados, ts):
    """resultados = [{"nombre":..., "tipo":..., "filas":[...]}, ...]"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    borde = Border(**{s:Side(style="thin",color="AAAAAA") for s in ["left","right","top","bottom"]})

    def estilo(cell, bold=False, bg=None, fg="000000", sz=10, ha="left", wrap=True):
        cell.font = Font(name="Arial", bold=bold, size=sz, color=fg)
        cell.alignment = Alignment(horizontal=ha, vertical="top", wrap_text=wrap)
        cell.border = borde
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    # ── Hoja Resumen ──────────────────────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Resumen"
    ws0.merge_cells("A1:D1")
    c = ws0.cell(row=1, column=1, value="REGISTRO MAESTRO CONSOLIDADO — ANÁLISIS DE ACTAS")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borde; ws0.row_dimensions[1].height = 28

    for col, enc in enumerate(["ARCHIVO","TIPO DE DOCUMENTO","FECHA PROCESAMIENTO","HOJA DE DETALLE"], 1):
        cell = ws0.cell(row=2, column=col)
        cell.value = enc
        estilo(cell, bold=True, bg="1F3864", fg="FFFFFF", sz=10, ha="center")
    ws0.row_dimensions[2].height = 20

    for i, res in enumerate(resultados):
        fila = i + 3
        bg = "F2F2F2" if i % 2 == 0 else "FFFFFF"
        vals = [res["nombre"], res["tipo"], datetime.now().strftime("%d/%m/%Y %H:%M"), f"Ver hoja: {res['hoja']}"]
        for col, val in enumerate(vals, 1):
            cell = ws0.cell(row=fila, column=col, value=val)
            estilo(cell, bg=bg, sz=10)
        ws0.row_dimensions[fila].height = 22

    ws0.column_dimensions["A"].width = 45
    ws0.column_dimensions["B"].width = 35
    ws0.column_dimensions["C"].width = 22
    ws0.column_dimensions["D"].width = 22
    ws0.freeze_panes = "A3"

    # ── Una hoja por archivo ──────────────────────────────────────────────────
    for res in resultados:
        nombre_hoja = res["hoja"]
        ws = wb.create_sheet(title=nombre_hoja)

        ws.merge_cells("A1:C1")
        c = ws.cell(row=1, column=1, value=f"REGISTRO MAESTRO — {res['nombre'].upper()}")
        c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde; ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:C2")
        c2 = ws.cell(row=2, column=1, value=f"Tipo: {res['tipo']}  |  Procesado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        c2.font = Font(name="Arial", size=9, color="FFFFFF", italic=True)
        c2.fill = PatternFill("solid", fgColor="2E75B6")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = borde; ws.row_dimensions[2].height = 16

        for col, enc in enumerate(["CATEGORÍA","DATO ACTUALIZADO / CLÁUSULA","OBSERVACIONES / CAMBIOS CLAVE"], 1):
            cell = ws.cell(row=3, column=col, value=enc)
            estilo(cell, bold=True, bg="1F3864", fg="FFFFFF", sz=10, ha="center")
        ws.row_dimensions[3].height = 20

        for j, f in enumerate(res["filas"]):
            row = j + 4
            bg = ("FFF2CC" if "Capital" in f["cat"] else
                  "FFE0B2" if "REPSE"  in f["cat"] else
                  "E2EFDA" if any(x in f["cat"] for x in ["Administración","Poderes","Facultades"]) else
                  "F2F2F2" if j%2==0 else "FFFFFF")
            for col, val in enumerate([f["cat"], f["dato"], f["obs"]], 1):
                cell = ws.cell(row=row, column=col, value=val)
                estilo(cell, bold=(col==1), bg=bg, sz=10)
            ws.row_dimensions[row].height = 75

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 60
        ws.freeze_panes = "A4"

    ruta = OUTPUT_DIR / f"RegistroMaestro_Lote_{ts}.xlsx"
    wb.save(str(ruta))
    return ruta

def guardar_csv_lote(resultados, ts):
    ruta = OUTPUT_DIR / f"RegistroMaestro_Lote_{ts}.csv"
    with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Fecha","Archivo","Tipo","Categoría","Dato Actualizado / Cláusula","Observaciones"])
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
        for res in resultados:
            for row in res["filas"]:
                w.writerow([fecha, res["nombre"], res["tipo"], row["cat"], row["dato"], row["obs"]])
    return ruta

# ── HTML ──────────────────────────────────────────────────────────────────────
def html_principal(api_key_preconfigurada):
    mostrar_api = "false" if api_key_preconfigurada else "true"
    return """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Analizador Legal — Actas Corporativas MX</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;min-height:100vh}
  header{background:linear-gradient(135deg,#1F3864,#2E75B6);color:#fff;padding:24px 40px;box-shadow:0 2px 8px rgba(0,0,0,.3)}
  header h1{font-size:1.6rem}
  header p{margin-top:5px;font-size:.88rem;opacity:.85}
  .badge{display:inline-block;background:#FFF2CC;color:#7B6000;border-radius:20px;padding:3px 12px;font-size:.76rem;font-weight:600;margin-top:7px}
  .logout{float:right;color:rgba(255,255,255,.7);font-size:.82rem;text-decoration:none;margin-top:4px}
  .logout:hover{color:#fff}
  .btn-cfg{float:right;clear:right;background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.35);color:#fff;padding:5px 14px;border-radius:20px;font-size:.8rem;cursor:pointer;margin-top:6px;transition:.2s}
  .btn-cfg:hover{background:rgba(255,255,255,.28)}
  .cfg-panel{display:none;background:#1a3260;border-top:1px solid rgba(255,255,255,.1);padding:16px 40px}
  .cfg-panel.open{display:flex;align-items:center;gap:12px;flex-wrap:wrap}
  .cfg-panel label{color:rgba(255,255,255,.85);font-size:.85rem;font-weight:600;white-space:nowrap}
  .cfg-panel input{flex:1;min-width:260px;padding:9px 14px;border-radius:8px;border:none;font-size:.9rem;font-family:monospace;background:rgba(255,255,255,.12);color:#fff}
  .cfg-panel input::placeholder{color:rgba(255,255,255,.4)}
  .cfg-panel input:focus{outline:none;background:rgba(255,255,255,.2)}
  .btn-save-cfg{padding:9px 22px;background:#2E75B6;color:#fff;border:none;border-radius:8px;font-weight:600;font-size:.88rem;cursor:pointer;white-space:nowrap}
  .btn-save-cfg:hover{background:#1F5fa6}
  .cfg-saved{color:#a8e6a3;font-size:.82rem;display:none;font-weight:600}
  main{max-width:980px;margin:32px auto;padding:0 20px}
  .card{background:#fff;border-radius:14px;box-shadow:0 2px 12px rgba(0,0,0,.08);margin-bottom:26px;overflow:hidden}
  .card-header{background:#1F3864;color:#fff;padding:15px 24px;font-size:.98rem;font-weight:600;display:flex;align-items:center;gap:8px}
  .card-body{padding:24px}

  /* Tabs */
  .tabs{display:flex;gap:0;border-bottom:2px solid #e0e8f5;margin-bottom:22px}
  .tab{padding:10px 24px;cursor:pointer;font-size:.92rem;font-weight:600;color:#888;border-bottom:3px solid transparent;margin-bottom:-2px;transition:.2s}
  .tab.active{color:#1F3864;border-bottom-color:#2E75B6}
  .tab-panel{display:none}.tab-panel.active{display:block}

  /* Drop zones */
  .drop-zone{border:2.5px dashed #2E75B6;border-radius:10px;padding:40px 24px;text-align:center;cursor:pointer;transition:.2s;background:#f7faff}
  .drop-zone:hover,.drop-zone.drag{border-color:#1F3864;background:#e8f0fb}
  .drop-zone .icon{font-size:2.6rem;margin-bottom:10px}
  .drop-zone p{color:#2E75B6;font-weight:600;font-size:1rem}
  .drop-zone small{color:#888;margin-top:5px;display:block}
  .file-list{margin-top:14px;max-height:160px;overflow-y:auto}
  .file-item{display:flex;align-items:center;gap:8px;padding:6px 10px;background:#f2f6fb;border-radius:6px;margin-bottom:5px;font-size:.85rem;color:#1F3864}
  .file-item .fi{font-size:1rem}

  /* API Key */
  .api-row{margin-top:16px}
  .api-row label{font-size:.85rem;color:#555;font-weight:600;display:block;margin-bottom:6px}
  .api-row input{width:100%;padding:11px 16px;border:1.5px solid #ccd;border-radius:8px;font-size:.93rem;font-family:monospace}
  .api-row input:focus{outline:none;border-color:#2E75B6}
  .api-ok{background:#E2EFDA;color:#1a5e20;padding:10px 14px;border-radius:8px;font-size:.85rem;margin-top:16px;border-left:4px solid #4CAF50}

  button.primary{background:linear-gradient(135deg,#1F3864,#2E75B6);color:#fff;border:none;padding:13px;border-radius:8px;font-size:1rem;font-weight:600;cursor:pointer;transition:.2s;width:100%;margin-top:16px}
  button.primary:hover{opacity:.9;transform:translateY(-1px)}
  button.primary:disabled{opacity:.45;cursor:not-allowed;transform:none}

  /* Progress */
  .progress{display:none;margin-top:16px}
  .progress-bar{height:8px;background:#e0e8f5;border-radius:4px;overflow:hidden}
  .progress-fill{height:100%;width:0%;background:linear-gradient(90deg,#2E75B6,#1F3864);border-radius:4px;transition:width .5s}
  .progress-text{font-size:.84rem;color:#2E75B6;margin-top:7px;font-weight:500}
  .progress-counter{font-size:.8rem;color:#888;margin-top:3px}

  /* Results */
  .result-area{display:none}
  .alert{padding:12px 18px;border-radius:8px;margin-bottom:14px;font-size:.88rem}
  .alert.success{background:#E2EFDA;color:#1a5e20;border-left:4px solid #4CAF50}
  .alert.error{background:#FDECEA;color:#b71c1c;border-left:4px solid #f44336}
  .alert.warn{background:#FFF8E1;color:#7B6000;border-left:4px solid #FFC107}
  table{width:100%;border-collapse:collapse;font-size:.86rem;margin-top:10px}
  th{background:#1F3864;color:#fff;padding:10px 13px;text-align:left;font-size:.8rem;text-transform:uppercase;letter-spacing:.4px}
  td{padding:9px 13px;vertical-align:top;border-bottom:1px solid #e8eef5;line-height:1.5}
  tr:nth-child(even) td{background:#f7faff}
  td:first-child{font-weight:600;color:#1F3864;white-space:nowrap;min-width:150px}
  td.capital{background:#FFF9E6!important}
  td.repse{background:#FFF3E0!important}
  td.admin{background:#F1F8E9!important}
  .dl-bar{display:flex;gap:12px;margin-top:16px;flex-wrap:wrap}
  .dl-btn{flex:1;min-width:180px;padding:11px;border-radius:8px;border:2px solid #2E75B6;background:#fff;color:#2E75B6;font-weight:600;cursor:pointer;font-size:.88rem;text-align:center;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:6px;transition:.2s}
  .dl-btn:hover{background:#2E75B6;color:#fff}
  .batch-summary{background:#f0f4f8;border-radius:8px;padding:14px;margin-bottom:14px}
  .batch-summary h3{color:#1F3864;font-size:.95rem;margin-bottom:8px}
  .batch-item{display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid #e0e8f5;font-size:.84rem}
  .batch-item:last-child{border-bottom:none}
  .dot{width:8px;height:8px;border-radius:50%;flex-shrink:0}
  .dot.ok{background:#4CAF50}.dot.err{background:#f44336}

  /* Steps */
  .steps{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
  .step{text-align:center;padding:14px 10px}
  .step .n{width:34px;height:34px;border-radius:50%;background:#2E75B6;color:#fff;font-weight:700;font-size:.95rem;display:flex;align-items:center;justify-content:center;margin:0 auto 9px}
  .step p{font-size:.83rem;color:#555;line-height:1.5}
  .tag{display:inline-block;background:#e8f0fb;color:#1F3864;border-radius:4px;padding:2px 8px;font-size:.75rem;font-weight:600;margin:2px}
  footer{text-align:center;padding:24px;color:#aaa;font-size:.78rem}
</style>
</head>
<body>
<header>
  <a class="logout" href="/logout">Cerrar sesión</a>
  <button class="btn-cfg" onclick="toggleConfig()">⚙️ Configuración</button>
  <h1>⚖️ Analizador Legal de Actas Corporativas</h1>
  <p>Derecho Corporativo Mexicano — Registro Maestro automático</p>
  <span class="badge" id="apiStatus">🔑 API Key no configurada</span>
</header>
<div class="cfg-panel" id="cfgPanel">
  <label>🔑 Google Gemini API Key</label>
  <input type="password" id="cfgApiKey" placeholder="AIzaSy-XXXXXXXXXXXXXXXX" oninput="document.getElementById('cfgSaved').style.display='none'">
  <button class="btn-save-cfg" onclick="guardarApiKey()">💾 Guardar</button>
  <span class="cfg-saved" id="cfgSaved">✅ ¡Guardada!</span>
</div>

<main>
  <div class="card">
    <div class="card-header">📋 ¿Cómo funciona?</div>
    <div class="card-body">
      <div class="steps">
        <div class="step"><div class="n">1</div><p>Sube <strong>un acta</strong> o selecciona una <strong>carpeta completa</strong> de documentos</p></div>
        <div class="step"><div class="n">2</div><p>Claude analiza cada documento como <strong>experto en Derecho Corporativo MX</strong></p></div>
        <div class="step"><div class="n">3</div><p>Descarga el <strong>Registro Maestro</strong> consolidado en Excel y CSV</p></div>
      </div>
      <p style="text-align:center;margin-top:10px;font-size:.8rem;color:#888">
        <span class="tag">Actas Constitutivas</span><span class="tag">Asambleas Ordinarias</span><span class="tag">Asambleas Extraordinarias</span><span class="tag">Sesiones de Consejo</span><span class="tag">PDFs escaneados</span><span class="tag">DOCX</span>
      </p>
    </div>
  </div>

  <div class="card">
    <div class="card-header">📤 Subir Documentos</div>
    <div class="card-body">
      <div class="tabs">
        <div class="tab active" onclick="switchTab('single')">📄 Un archivo</div>
        <div class="tab" onclick="switchTab('batch')">📁 Carpeta completa</div>
      </div>

      <!-- Tab: Un archivo -->
      <div class="tab-panel active" id="panel-single">
        <div class="drop-zone" id="dropSingle" onclick="document.getElementById('inputSingle').click()">
          <div class="icon">📄</div>
          <p>Arrastra tu acta aquí o haz clic para seleccionar</p>
          <small>PDF o DOCX · Máximo 50 MB · Hasta 30 páginas</small>
        </div>
        <input type="file" id="inputSingle" accept=".pdf,.docx" style="display:none" onchange="onSingleFile(this)">
        <div id="single-name" style="display:none;margin-top:12px;font-size:.88rem;color:#1F3864;font-weight:600"></div>
      </div>

      <!-- Tab: Carpeta -->
      <div class="tab-panel" id="panel-batch">
        <div class="drop-zone" id="dropBatch" onclick="document.getElementById('inputBatch').click()" id="dropBatch"
             ondragover="event.preventDefault();this.classList.add('drag')"
             ondragleave="this.classList.remove('drag')"
             ondrop="onDropCarpeta(event)">
          <div class="icon">📁</div>
          <p>Haz clic para seleccionar una carpeta completa</p>
          <small>Se detectarán automáticamente todos los PDF y DOCX dentro de la carpeta</small>
        </div>
        <input type="file" id="inputBatch" webkitdirectory mozdirectory directory multiple style="display:none" onchange="onBatchFiles(this)">
        <div id="batchCount" style="display:none;margin-top:12px;padding:10px 14px;background:#e8f0fb;border-radius:8px;font-size:.88rem;color:#1F3864;font-weight:600"></div>
        <div class="file-list" id="batchList"></div>
      </div>

      <!-- API Key (oculta, se maneja desde Configuración) -->
      """ + ("""<div class="api-ok">✅ API Key configurada en el servidor.</div>""" if api_key_preconfigurada else """
      <div id="apiWarning" style="display:none;padding:10px 14px;background:#FFF8E1;border-left:4px solid #FFC107;border-radius:8px;font-size:.85rem;color:#7B6000;margin-top:14px">
        ⚠️ Configura tu API Key con el botón <strong>⚙️ Configuración</strong> (arriba a la derecha) antes de analizar.
      </div>""") + """

      <button class="primary" id="btnAnalizar" onclick="analizar()" disabled>🔍 Analizar Documento(s)</button>

      <div class="progress" id="progressBox">
        <div class="progress-bar"><div class="progress-fill" id="pFill"></div></div>
        <div class="progress-text" id="pText">Iniciando...</div>
        <div class="progress-counter" id="pCounter"></div>
      </div>
    </div>
  </div>

  <div class="card result-area" id="resultArea">
    <div class="card-header">✅ Registro Maestro Generado</div>
    <div class="card-body">
      <div id="alertBox"></div>
      <div id="batchSummary"></div>
      <table id="resultTable" style="display:none">
        <thead><tr><th>Categoría</th><th>Dato Actualizado / Cláusula</th><th>Observaciones / Cambios Clave</th></tr></thead>
        <tbody id="resultBody"></tbody>
      </table>
      <div class="dl-bar" id="dlBar" style="display:none">
        <a class="dl-btn" id="dlExcel" href="#" download>📊 Descargar Excel (.xlsx)</a>
        <a class="dl-btn" id="dlCsv" href="#" download>📋 Descargar CSV</a>
      </div>
    </div>
  </div>
</main>
<footer>Uso interno · Documentos procesados de forma segura · © """ + str(datetime.now().year) + """</footer>

<script>
const API_KEY_ENV = """ + mostrar_api + """;
let modo = 'single';
let archivoSingle = null;
let archivosBatch = [];

// ── Config: guardar/leer API Key en localStorage ───────────────────────────
function toggleConfig() {
  const panel = document.getElementById('cfgPanel');
  panel.classList.toggle('open');
  if (panel.classList.contains('open')) {
    const saved = localStorage.getItem('gemini_api_key') || '';
    document.getElementById('cfgApiKey').value = saved;
  }
}
function guardarApiKey() {
  const key = document.getElementById('cfgApiKey').value.trim();
  if (!key) return;
  localStorage.setItem('gemini_api_key', key);
  document.getElementById('cfgSaved').style.display = 'inline';
  actualizarBadge(key);
  checkReady();
  setTimeout(() => document.getElementById('cfgPanel').classList.remove('open'), 1200);
}
function actualizarBadge(key) {
  const badge = document.getElementById('apiStatus');
  const warn  = document.getElementById('apiWarning');
  if (key) {
    badge.textContent = '✅ API Key configurada';
    badge.style.background = '#E2EFDA';
    badge.style.color = '#1a5e20';
    if (warn) warn.style.display = 'none';
  } else {
    badge.textContent = '🔑 API Key no configurada';
    badge.style.background = '#FFF2CC';
    badge.style.color = '#7B6000';
    if (warn) warn.style.display = 'block';
  }
}
// Al cargar la página, restaurar API Key guardada
window.addEventListener('load', () => {
  const saved = localStorage.getItem('gemini_api_key') || '';
  actualizarBadge(saved);
  checkReady();
});

function switchTab(t) {
  modo = t;
  document.querySelectorAll('.tab').forEach((el,i) => el.classList.toggle('active', (i===0&&t==='single')||(i===1&&t==='batch')));
  document.getElementById('panel-single').classList.toggle('active', t==='single');
  document.getElementById('panel-batch').classList.toggle('active', t==='batch');
  checkReady();
}

// Drop para archivo único
const dz = document.getElementById('dropSingle');
dz.addEventListener('dragover', e => { e.preventDefault(); dz.classList.add('drag'); });
dz.addEventListener('dragleave', () => dz.classList.remove('drag'));
dz.addEventListener('drop', e => { e.preventDefault(); dz.classList.remove('drag'); const f=e.dataTransfer.files[0]; if(f) setSingle(f); });

function onSingleFile(inp) { if(inp.files[0]) setSingle(inp.files[0]); }
function setSingle(f) {
  archivoSingle = f;
  const el = document.getElementById('single-name');
  el.textContent = '📎 ' + f.name + ' (' + (f.size/1024/1024).toFixed(2) + ' MB)';
  el.style.display = 'block';
  checkReady();
}

function onBatchFiles(inp) {
  // Filtra solo PDF y DOCX de todos los archivos de la carpeta seleccionada
  archivosBatch = Array.from(inp.files).filter(f => /\.(pdf|docx)$/i.test(f.name));
  renderBatchList();
  checkReady();
}

function onDropCarpeta(event) {
  event.preventDefault();
  document.getElementById('dropBatch').classList.remove('drag');
  // Drag & drop de archivos individuales (fallback)
  const files = Array.from(event.dataTransfer.files).filter(f => /\.(pdf|docx)$/i.test(f.name));
  if (files.length > 0) { archivosBatch = files; renderBatchList(); checkReady(); }
}

function renderBatchList() {
  const countEl = document.getElementById('batchCount');
  const listEl  = document.getElementById('batchList');
  if (archivosBatch.length === 0) {
    countEl.style.display = 'none'; listEl.innerHTML = ''; return;
  }
  const totalMB = archivosBatch.reduce((s,f) => s + f.size/1024/1024, 0).toFixed(2);
  countEl.style.display = 'block';
  countEl.innerHTML = `📁 ${archivosBatch.length} archivos encontrados en la carpeta &nbsp;·&nbsp; ${totalMB} MB en total`;
  listEl.innerHTML = archivosBatch.map(f =>
    `<div class="file-item"><span class="fi">${f.name.endsWith('.pdf')?'📄':'📝'}</span>` +
    `<span>${f.name}</span>` +
    `<span style="margin-left:auto;color:#888;font-size:.8rem">${(f.size/1024/1024).toFixed(2)} MB</span></div>`
  ).join('');
}

function getApiKey() {
  if (API_KEY_ENV) return '';  // el servidor usa la variable de entorno
  return localStorage.getItem('gemini_api_key') || '';
}

function checkReady() {
  const tieneArchivos = modo==='single' ? !!archivoSingle : archivosBatch.length > 0;
  const tieneKey = API_KEY_ENV || getApiKey().length > 20;
  const warn = document.getElementById('apiWarning');
  if (warn) warn.style.display = (!tieneKey) ? 'block' : 'none';
  document.getElementById('btnAnalizar').disabled = !(tieneArchivos && tieneKey);
}

function setProgress(pct, txt, counter='') {
  document.getElementById('pFill').style.width = pct + '%';
  document.getElementById('pText').textContent = txt;
  document.getElementById('pCounter').textContent = counter;
}

async function analizar() {
  const btn = document.getElementById('btnAnalizar');
  btn.disabled = true;
  document.getElementById('progressBox').style.display = 'block';
  document.getElementById('resultArea').style.display = 'none';

  if (modo === 'single') {
    await analizarSingle();
  } else {
    await analizarLote();
  }
  btn.disabled = false;
}

async function analizarSingle() {
  const pasos = [[20,'📄 Leyendo el documento...'],[45,'🖼️ Procesando páginas...'],[70,'🤖 Claude analizando las cláusulas...'],[90,'📊 Estructurando el Registro Maestro...']];
  let pi = 0;
  const tick = setInterval(() => { if(pi<pasos.length){setProgress(...pasos[pi]);pi++;} }, 4500);
  try {
    const fd = new FormData();
    fd.append('file', archivoSingle);
    if (!API_KEY_ENV) fd.append('api_key', getApiKey());
    const res = await fetch('/analizar', {method:'POST',body:fd});
    const data = await res.json();
    clearInterval(tick); setProgress(100,'✅ ¡Listo!');
    if (data.error) { showAlert('error','❌ '+data.error); }
    else {
      renderTabla(data.filas);
      if (data.excel_url) {
        document.getElementById('dlExcel').href = data.excel_url;
        document.getElementById('dlCsv').href   = data.csv_url;
        document.getElementById('dlBar').style.display = 'flex';
      }
    }
  } catch(e) { clearInterval(tick); showAlert('error','❌ Error: '+e.message); }
  document.getElementById('resultArea').style.display = 'block';
}

async function analizarLote() {
  const total = archivosBatch.length;
  const resultados = [];
  let exitosos = 0, fallidos = 0;

  for (let i = 0; i < total; i++) {
    const f = archivosBatch[i];
    setProgress(Math.round((i/total)*85), `🤖 Analizando: ${f.name}`, `Archivo ${i+1} de ${total}`);
    try {
      const fd = new FormData();
      fd.append('file', f);
      if (!API_KEY_ENV) fd.append('api_key', getApiKey());
      const res = await fetch('/analizar', {method:'POST',body:fd});
      const data = await res.json();
      if (data.error) { resultados.push({nombre:f.name,ok:false,error:data.error}); fallidos++; }
      else { resultados.push({nombre:f.name,ok:true,tipo:data.tipo,filas:data.filas}); exitosos++; }
    } catch(e) { resultados.push({nombre:f.name,ok:false,error:e.message}); fallidos++; }
  }

  setProgress(95,'📊 Generando Registro Maestro consolidado...');
  try {
    const fd2 = new FormData();
    fd2.append('resultados', JSON.stringify(resultados.filter(r=>r.ok)));
    if (!API_KEY_ENV) fd2.append('api_key', getApiKey());
    const res2 = await fetch('/consolidar', {method:'POST',body:fd2});
    const data2 = await res2.json();
    setProgress(100,'✅ ¡Lote completo!', `${exitosos} exitosos · ${fallidos} con error`);
    renderResumenLote(resultados);
    if (data2.excel_url) {
      document.getElementById('dlExcel').href = data2.excel_url;
      document.getElementById('dlCsv').href   = data2.csv_url;
      document.getElementById('dlBar').style.display = 'flex';
    }
  } catch(e) { setProgress(100,'⚠️ Error al consolidar',''); showAlert('error','❌ '+e.message); }
  document.getElementById('resultArea').style.display = 'block';
}

function renderResumenLote(resultados) {
  document.getElementById('batchSummary').innerHTML =
    `<div class="batch-summary"><h3>📋 Resumen del procesamiento</h3>` +
    resultados.map(r => `<div class="batch-item"><div class="dot ${r.ok?'ok':'err'}"></div><span>${r.nombre}</span><span style="margin-left:auto;font-size:.8rem;color:${r.ok?'#1a5e20':'#b71c1c'}">${r.ok?r.tipo:'Error: '+r.error}</span></div>`).join('') +
    `</div>`;
  document.getElementById('resultTable').style.display = 'none';
}

function renderTabla(filas) {
  const tbody = document.getElementById('resultBody');
  tbody.innerHTML = '';
  document.getElementById('alertBox').innerHTML = '';
  document.getElementById('batchSummary').innerHTML = '';
  document.getElementById('resultTable').style.display = 'table';
  filas.forEach(f => {
    const tr = document.createElement('tr');
    const cls = f.cat.includes('Capital')?'capital':f.cat.includes('REPSE')?'repse':(f.cat.includes('Administración')||f.cat.includes('Facultad')||f.cat.includes('Poder'))?'admin':'';
    tr.innerHTML = `<td class="${cls}">${f.cat}</td><td class="${cls}">${f.dato.replace(/\\n/g,'<br>')}</td><td class="${cls}">${f.obs.replace(/\\n/g,'<br>')}</td>`;
    tbody.appendChild(tr);
  });
}

function showAlert(type, msg) {
  document.getElementById('alertBox').innerHTML = `<div class="alert ${type}">${msg}</div>`;
}
</script>
</body>
</html>"""

HTML_LOGIN = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Acceso — Analizador Legal MX</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:'Segoe UI',Arial,sans-serif;background:linear-gradient(135deg,#1F3864,#2E75B6);min-height:100vh;display:flex;align-items:center;justify-content:center}
  .box{background:#fff;border-radius:16px;box-shadow:0 8px 32px rgba(0,0,0,.25);padding:44px 40px;width:100%;max-width:400px;text-align:center}
  .logo{font-size:2.8rem;margin-bottom:10px}
  h1{color:#1F3864;font-size:1.3rem;margin-bottom:6px}
  p{color:#888;font-size:.88rem;margin-bottom:28px}
  input{width:100%;padding:13px 16px;border:1.5px solid #dde;border-radius:8px;font-size:1rem;font-family:monospace;margin-bottom:16px;outline:none;transition:.2s}
  input:focus{border-color:#2E75B6}
  button{width:100%;padding:13px;background:linear-gradient(135deg,#1F3864,#2E75B6);color:#fff;border:none;border-radius:8px;font-size:1rem;font-weight:600;cursor:pointer}
  button:hover{opacity:.9}
  .error{background:#FDECEA;color:#b71c1c;border-radius:8px;padding:10px 14px;font-size:.88rem;margin-bottom:16px;text-align:left}
  footer{margin-top:22px;font-size:.76rem;color:#bbb}
</style>
</head>
<body>
<div class="box">
  <div class="logo">⚖️</div>
  <h1>Analizador Legal de Actas</h1>
  <p>Acceso restringido — solo personal autorizado</p>
  {% if error %}<div class="error">❌ Contraseña incorrecta.</div>{% endif %}
  <form method="POST" action="/login">
    <input type="password" name="password" placeholder="Contraseña de acceso" autofocus required>
    <button type="submit">Entrar →</button>
  </form>
  <footer>Powered by Claude (Anthropic)</footer>
</div>
</body>
</html>"""

# ── Rutas ─────────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        if request.form.get("password","") == APP_PASSWORD:
            session["autenticado"] = True
            return redirect(url_for("index"))
        return render_template_string(HTML_LOGIN, error=True)
    return render_template_string(HTML_LOGIN, error=False)

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

@app.route("/")
@login_required
def index():
    return html_principal(bool(API_KEY_ENV))

@app.route("/analizar", methods=["POST"])
@login_required
def analizar_endpoint():
    try:
        api_key = API_KEY_ENV or request.form.get("api_key","").strip()
        if not api_key:
            return jsonify({"error":"Ingresa tu API Key de Google Gemini para continuar."}), 400
        f = request.files.get("file")
        if not f:
            return jsonify({"error":"No se recibió ningún archivo."}), 400
        nombre = f.filename
        ruta   = UPLOAD_DIR / nombre
        f.save(str(ruta))
        filas, tipo = procesar_archivo(ruta, api_key)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = guardar_excel_lote([{"nombre":nombre,"tipo":tipo,"filas":filas,
                                          "hoja":nombre[:28].replace("/","_")}], ts)
        csv_path   = guardar_csv_lote([{"nombre":nombre,"tipo":tipo,"filas":filas}], ts)
        return jsonify({"filas":filas,"tipo":tipo,
                        "excel_url":f"/descargar/{excel_path.name}",
                        "csv_url":  f"/descargar/{csv_path.name}"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/consolidar", methods=["POST"])
@login_required
def consolidar_endpoint():
    try:
        import json
        datos = json.loads(request.form.get("resultados","[]"))
        if not datos:
            return jsonify({"error":"No hay documentos procesados para consolidar."}), 400
        resultados = []
        for i, d in enumerate(datos):
            hoja = d["nombre"][:28].replace("/","_").replace(".","_")
            resultados.append({"nombre":d["nombre"],"tipo":d["tipo"],
                                "filas":d["filas"],"hoja":hoja})
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = guardar_excel_lote(resultados, ts)
        csv_path   = guardar_csv_lote(resultados, ts)
        return jsonify({"excel_url":f"/descargar/{excel_path.name}",
                        "csv_url":  f"/descargar/{csv_path.name}"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/descargar/<nombre>")
@login_required
def descargar(nombre):
    ruta = OUTPUT_DIR / nombre
    if ruta.exists():
        return send_file(str(ruta), as_attachment=True)
    return "Archivo no encontrado", 404

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") == "development"
    print(f"\n{'='*50}\n  ⚖️  ANALIZADOR LEGAL — WEB\n{'='*50}")
    print(f"  🌐 http://localhost:{port}")
    print(f"  🔑 API Key env: {'✅ configurada' if API_KEY_ENV else '❌ no configurada (se pide en la web)'}")
    print(f"  🔒 Password:    {'✅ activa' if APP_PASSWORD else '❌ sin login'}")
    print(f"{'='*50}\n")
    app.run(debug=debug, host="0.0.0.0", port=port)
